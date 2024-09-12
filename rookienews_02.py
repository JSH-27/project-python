

import pandas as pd  # pip install pandas 필요 (대략 1분 소요)


import collections
data = ['대한','민국','우리','나라','우리','우리','나라','민국','우리','나라','우리','우리','나라','대한민국','우리나라','민국','우리','나라','우리','우리','나라','민국','우리','나라','우리','우리','나라','대한민국','우리나라','민국','우리','나라','우리','우리','나라','민국','우리','나라','우리','우리','나라','대한민국','우리나라','민국','우리','나라','우리','우리','나라','민국','우리','나라','우리','우리','나라','대한민국','우리나라']
frequency = collections.Counter(data)
for key, value in frequency.items():
    print(f"key : {key}, value : {value}")
    print(type(key))
    print(type(value))
    print(f"최대 점수 : {max(value)}")
    print(f"최소 점수 : {min(value)}")
    # key_word = frequency[key].max(value)
    # print(key_word)
    # data2 = key.sort_values(by=value, ascending=False, inplace=True)
    # print(data2)

# ===================================================
# 엑셀파일 생성
from openpyxl import Workbook
import datetime, time

wb = Workbook()
ws = wb.active
CURRENT_DAY = datetime.datetime.now().strftime("%Y-%m-%d")
count_num = 4
# count_num = count_num + len(data)
ws.title = f'{CURRENT_DAY} Rookienews'

ws['C2'] = f'{CURRENT_DAY}의 주요뉴스들 Hot 키워드 / 헤드라인'
ws['C4'] = 'Hot 키워드'
ws['D4'] = '빈도수'
ws['F4'] = 'HEADLINE'
ws['G4'] = 'Link'
# ws.cell(row=5, column=2, value="")
# ws.cell(row=5, column=3, value="key_word")
# ws.cell(row=5, column=4, value="key_word_count")
# ws.cell(row=count_num, column=6, value=article_title)
# ws.cell(row=count_num, column=7, value=article_link)

wb.save(f"{CURRENT_DAY} Rookienews.xlsx")

# ===================================================